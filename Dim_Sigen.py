from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()

def autofit(ws, width=16, max_col=20):
    for col in range(1, max_col+1):
        ws.column_dimensions[get_column_letter(col)].width = width

# ========== SHEET 1: CATALOGUE ==========
ws_cat = wb.active
ws_cat.title = "Catalogue"

# Panneaux section
ws_cat.append(["Panneaux"])
ws_cat.append(["ID","P_STC_W","Voc","Vmp","Isc","Imp","alpha_V_%/°C"])
panels = [
    ["Trina450",450,52.9,44.6,10.74,10.09,-0.24],
    ["Trina505_DEG18",505,51.7,43.7,12.13,11.56,-0.25],
    ["Solux415",415,37.95,31.83,13.77,13.04,-0.28],
    ["Solux420",420,38.14,32.02,13.85,13.12,-0.28],
    ["Solux425",425,38.32,32.20,13.93,13.20,-0.28],
    ["TrinaS+475",475,39.0,36.9,14.72,11.94,-0.24],
    ["TrinaS+480",480,39.2,37.2,14.77,11.98,-0.24],
    ["TrinaS+485",485,39.4,37.4,14.84,12.06,-0.24],
    ["TrinaS+490",490,39.6,37.7,14.91,12.11,-0.24],
    ["TrinaS+495",495,39.8,38.0,14.97,12.15,-0.24],
    ["TrinaS+500",500,40.1,38.3,15.03,12.18,-0.24],
    ["TrinaS+505",505,40.3,38.5,15.09,12.22,-0.24],
]
first_panel_row = ws_cat.max_row + 1
for p in panels:
    ws_cat.append(p)
last_panel_row = ws_cat.max_row

# Onduleurs section
ws_cat.append([""])
ws_cat.append(["Onduleurs"])
ws_cat.append(["ID","P_AC_nom","P_DC_max","V_MPP_min","V_MPP_max",
               "V_DC_max","I_MPPT","Nb_MPPT","Type_reseau"])
inverters = [
    ("Sigen2.0",2000,4000,50,550,600,16,2,"Mono"),
    ("Sigen3.0",3000,6000,50,550,600,16,2,"Mono"),
    ("Sigen3.6",3680,7360,50,550,600,16,2,"Mono"),
    ("Sigen4.0",4000,8000,50,550,600,16,2,"Mono"),
    ("Sigen4.6",4600,9200,50,550,600,16,2,"Mono"),
    ("Sigen5.0",5000,10000,50,550,600,16,2,"Mono"),
    ("Sigen6.0",6000,12000,50,550,600,16,2,"Mono"),
    ("Sigen3T",3000,6000,160,1000,1100,16,2,"Tri 3x400"),
    ("Sigen4T",4000,8000,160,1000,1100,16,2,"Tri 3x400"),
    ("Sigen5T",5000,10000,160,1000,1100,16,2,"Tri 3x400"),
    ("Sigen6T",6000,12000,160,1000,1100,16,2,"Tri 3x400"),
    ("Sigen8T",8000,16000,160,1000,1100,32,2,"Tri 3x400"),
    ("Sigen10T",10000,20000,160,1000,1100,32,2,"Tri 3x400"),
    ("Sigen12T",12000,24000,160,1000,1100,32,2,"Tri 3x400"),
]
first_inv_row = ws_cat.max_row + 1
for inv in inverters:
    ws_cat.append(list(inv))
last_inv_row = ws_cat.max_row

# Batteries section
ws_cat.append([""])
ws_cat.append(["Batteries"])
ws_cat.append(["ID","Cap_kWh"])
bat_first_row = ws_cat.max_row + 1
bats = [["Sigen6",6],["Sigen10",10]]
for b in bats:
    ws_cat.append(b)
bat_last_row = ws_cat.max_row

autofit(ws_cat, max_col=10)

# ========== SHEET 2: CHOIX ==========
ws_choice = wb.create_sheet("Choix")

ws_choice["A1"]="Panneau"
ws_choice["A2"]="Nombre modules"
ws_choice["A3"]="Type réseau"
ws_choice["A4"]="Batterie ?"
ws_choice["A5"]="Capacité batterie (kWh)"
ws_choice["A6"]="Ratio DC/AC max"

ws_choice["B2"]=10
ws_choice["B3"]="Mono"
ws_choice["B4"]="Non"
ws_choice["B5"]=6
ws_choice["B6"]=1.3

# Data validations
dv_panel = DataValidation(type="list",
                          formula1=f"=Catalogue!$A${first_panel_row}:$A${last_panel_row}")
dv_net   = DataValidation(type="list", formula1="\"Mono,Tri 3x400\"")
dv_yesno = DataValidation(type="list", formula1="\"Oui,Non\"")

ws_choice.add_data_validation(dv_panel)
ws_choice.add_data_validation(dv_net)
ws_choice.add_data_validation(dv_yesno)

dv_panel.add(ws_choice["B1"])
dv_net.add(ws_choice["B3"])
dv_yesno.add(ws_choice["B4"])

# P_STC, P_DC
ws_choice["A8"]="P_STC panneau (W)"
ws_choice["B8"]=f"=IFERROR(VLOOKUP(B1,Catalogue!$A${first_panel_row}:$G${last_panel_row},2,FALSE),\"\")"
ws_choice["A9"]="Puissance DC totale (W)"
ws_choice["B9"]="=IF(B8<>\"\",B8*B2,\"\")"

# Inverter table
ws_choice.append([""])
start_row = 12
headers = ["Onduleur","P_AC","Type réseau","DC/AC","Type OK",
           "Ratio OK","Global OK","Batterie"]
for c,h in enumerate(headers,1):
    ws_choice.cell(start_row,c).value = h

row = start_row + 1
for offset in range(len(inverters)):
    cat_row = first_inv_row + offset
    ws_choice[f"A{row}"]=f"=Catalogue!A{cat_row}"
    ws_choice[f"B{row}"]=f"=Catalogue!B{cat_row}"
    ws_choice[f"C{row}"]=f"=Catalogue!I{cat_row}"
    ws_choice[f"D{row}"]=f"=IF($B$9<>\"\",$B$9/B{row},\"\")"
    ws_choice[f"E{row}"]=f"=IF(C{row}=$B$3,\"OK\",\"NO\")"
    ws_choice[f"F{row}"]=f"=IF(D{row}=\"\",\"\",IF(D{row}<=$B$6,\"OK\",\"NO\"))"
    ws_choice[f"G{row}"]=f"=IF(AND(E{row}=\"OK\",F{row}=\"OK\"),\"OK\",\"NO\")"
    ws_choice[f"H{row}"]=f"=IF($B$4<>\"Oui\",\"Aucune\",IF($B$5<=6,\"Sigen6\",\"Sigen10\"))"
    row+=1

autofit(ws_choice, max_col=8)

# ========== SHEET 3: PROFIL (Mensuel) ==========
ws_prof = wb.create_sheet("Profil")

ws_prof["A1"]="Conso annuelle (kWh)"
ws_prof["B1"]=3500
ws_prof["A2"]="Profil conso"
ws_prof["B2"]="Standard"

dv_prof = DataValidation(type="list", formula1="\"Standard,Hiver fort,Été fort\"")
ws_prof.add_data_validation(dv_prof)
dv_prof.add(ws_prof["B2"])

ws_prof.append([""])
ws_prof.append(["Mois","%_conso","Conso_kWh","Prod_PV_kWh","kWh_kWp_BEL","Autocons_kWh"])

months = ["Jan","Fév","Mar","Avr","Mai","Juin","Juil","Août","Sep","Oct","Nov","Déc"]
percent_std    = [8,8,8,8,8,8,8,8,8,8,8,8]
percent_winter = [10,10,10,9,8,7,6,6,7,8,9,10]
percent_summer = [6,6,7,8,9,10,11,11,10,8,7,7]

annual_kwh_kwp = 1034.0
distribution = [3.8,5.1,8.7,11.5,12.1,11.8,11.9,10.8,9.7,7.0,4.3,3.3]
kwh_kwp = [round(annual_kwh_kwp*d/100.0,1) for d in distribution]

start = 5
for i,m in enumerate(months):
    r = start + i
    ws_prof.cell(r,1).value = m
    ws_prof.cell(r,2).value = f"=CHOOSE(MATCH($B$2,{{\"Standard\",\"Hiver fort\",\"Été fort\"}},0),{percent_std[i]},{percent_winter[i]},{percent_summer[i]})"
    ws_prof.cell(r,3).value = f"=($B$1*B{r}/100)"
    ws_prof.cell(r,5).value = kwh_kwp[i]
    ws_prof.cell(r,4).value = f"=E{r}*Choix!$B$9/1000"
    ws_prof.cell(r,6).value = f"=IF(OR(C{r}=\"\",D{r}=\"\"),\"\",MIN(C{r},D{r}))"

autofit(ws_prof, max_col=6)

# Graph: Prod vs Conso
bar = BarChart()
bar.title = "Production vs Consommation (Belgique)"
bar.y_axis.title = "kWh"
bar.x_axis.title = "Mois"

data_bar = Reference(ws_prof, min_col=3, max_col=4, min_row=4, max_row=4+len(months))
cats = Reference(ws_prof, min_col=1, min_row=5, max_row=5+len(months)-1)
bar.add_data(data_bar, titles_from_data=True)
bar.set_categories(cats)
bar.height = 12
bar.width = 20
ws_prof.add_chart(bar, "H4")

# ========== SHEET 4: STRINGS (tensions froid/chaud) ==========
ws_str = wb.create_sheet("Strings")

ws_str["A1"] = "Vérification des tensions de strings"
ws_str["A3"] = "Panneau"
ws_str["A4"] = "Onduleur"
ws_str["A5"] = "Température min (°C)"
ws_str["A6"] = "Température max (°C)"
ws_str["A7"] = "Modules en série (N_series)"

ws_str["B5"] = -10
ws_str["B6"] = 70
ws_str["B7"] = 10

# Data validations for panel and inverter
dv_p2 = DataValidation(type="list",
                       formula1=f"=Catalogue!$A${first_panel_row}:$A${last_panel_row}")
dv_i2 = DataValidation(type="list",
                       formula1=f"=Catalogue!$A${first_inv_row}:$A${last_inv_row}")
ws_str.add_data_validation(dv_p2)
ws_str.add_data_validation(dv_i2)
dv_p2.add(ws_str["B3"])
dv_i2.add(ws_str["B4"])

# Lookups
ws_str["A9"]="Voc_module (V)"
ws_str["B9"]=f"=IFERROR(VLOOKUP(B3,Catalogue!$A${first_panel_row}:$G${last_panel_row},3,FALSE),\"\")"
ws_str["A10"]="Vmp_module (V)"
ws_str["B10"]=f"=IFERROR(VLOOKUP(B3,Catalogue!$A${first_panel_row}:$G${last_panel_row},4,FALSE),\"\")"
ws_str["A11"]="alpha_V (%/°C)"
ws_str["B11"]=f"=IFERROR(VLOOKUP(B3,Catalogue!$A${first_panel_row}:$G${last_panel_row},7,FALSE),\"\")"

ws_str["A13"]="V_DC_max onduleur (V)"
ws_str["B13"]=f"=IFERROR(VLOOKUP(B4,Catalogue!$A${first_inv_row}:$I${last_inv_row},6,FALSE),\"\")"
ws_str["A14"]="V_MPP_min (V)"
ws_str["B14"]=f"=IFERROR(VLOOKUP(B4,Catalogue!$A${first_inv_row}:$I${last_inv_row},4,FALSE),\"\")"
ws_str["A15"]="V_MPP_max (V)"
ws_str["B15"]=f"=IFERROR(VLOOKUP(B4,Catalogue!$A${first_inv_row}:$I${last_inv_row},5,FALSE),\"\")"

# String calculations
ws_str["A17"]="Voc_string_froid (V)"
ws_str["B17"]="=IF(OR(B9=\"\",B11=\"\"),\"\",B7*B9*(1+B11/100*(B5-25)))"
ws_str["A18"]="Vmp_string_chaud (V)"
ws_str["B18"]="=IF(OR(B10=\"\",B11=\"\"),\"\",B7*B10*(1+B11/100*(B6-25)))"

ws_str["A20"]="Check Voc <= V_DC_max"
ws_str["B20"]="=IF(OR(B17=\"\",B13=\"\"),\"\",IF(B17<=B13,\"OK\",\"DÉPASSE\"))"
ws_str["A21"]="Check Vmp dans plage MPPT"
ws_str["B21"]="=IF(OR(B18=\"\",B14=\"\",B15=\"\"),\"\",IF(AND(B18>=B14,B18<=B15),\"OK\",\"HORS PLAGE\"))"

autofit(ws_str, max_col=4)

# ========== SHEET 5: HEURES (profil journalier) ==========
ws_h = wb.create_sheet("Heures")

ws_h["A1"]="Mois (1-12)"
ws_h["B1"]=1
ws_h["A2"]="Conso jour (kWh)"
ws_h["B2"]="=IF(OR($B$1<1,$B$1>12),\"\",INDEX(Profil!$C$5:$C$16,$B$1)/INDEX({31;28;31;30;31;30;31;31;30;31;30;31},$B$1))"
ws_h["A3"]="Prod jour (kWh)"
ws_h["B3"]="=IF(OR($B$1<1,$B$1>12),\"\",INDEX(Profil!$D$5:$D$16,$B$1)/INDEX({31;28;31;30;31;30;31;31;30;31;30;31},$B$1))"

ws_h.append([""])
ws_h.append(["Heure","Conso_h_kWh","Prod_h_kWh","Autocons_h_kWh"])

pv_frac = [0,0,0,0,0,0.01,0.04,0.07,0.10,0.13,0.14,0.14,0.13,0.10,0.07,0.04,0.02,0,0,0,0,0,0,0]

start = 6
for h in range(24):
    r = start + h
    ws_h.cell(r,1).value = h
    ws_h.cell(r,2).value = f"=IF($B$2=\"\",\"\",$B$2/24)"
    frac = pv_frac[h]
    ws_h.cell(r,3).value = f"=IF($B$3=\"\",\"\",$B$3*{frac})"
    ws_h.cell(r,4).value = f"=IF(OR(B{r}=\"\",C{r}=\"\"),\"\",MIN(B{r},C{r}))"

autofit(ws_h, max_col=5)

# ========== SHEET 6: SYNTHÈSE ==========
ws_syn = wb.create_sheet("Synthese")
ws_syn["A1"]="Synthèse client (Belgique)"

ws_syn["A3"]="Panneau"
ws_syn["B3"]="=Choix!B1"
ws_syn["A4"]="Modules"
ws_syn["B4"]="=Choix!B2"
ws_syn["A5"]="Puissance DC totale (W)"
ws_syn["B5"]="=Choix!B9"

ws_syn["A7"]="Onduleur recommandé"
ws_syn["B7"]="=INDEX(Choix!A:A,MATCH(\"OK\",Choix!G:G,0))"
ws_syn["A8"]="Puissance AC (W)"
ws_syn["B8"]="=INDEX(Choix!B:B,MATCH(\"OK\",Choix!G:G,0))"

ws_syn["A10"]="Conso annuelle (kWh)"
ws_syn["B10"]="=SUM(Profil!C5:C16)"
ws_syn["A11"]="Prod PV annuelle (kWh)"
ws_syn["B11"]="=SUM(Profil!D5:D16)"
ws_syn["A12"]="Autocons annuelle (kWh)"
ws_syn["B12"]="=SUM(Profil!F5:F16)"
ws_syn["A13"]="Taux autocons"
ws_syn["B13"]="=IF(B11=0,\"\",B12/B11)"
ws_syn["A14"]="Taux couverture"
ws_syn["B14"]="=IF(B10=0,\"\",B12/B10)"

ws_syn["A16"]="Batterie activée ?"
ws_syn["B16"]="=Choix!B4"
ws_syn["A17"]="Capacité batterie cible (kWh)"
ws_syn["B17"]="=Choix!B5"
ws_syn["A18"]="Modèle batterie"
ws_syn["B18"]="=IF(B16<>\"Oui\",\"Aucune\",IF(Choix!B5<=6,\"Sigen6\",\"Sigen10\"))"

autofit(ws_syn, max_col=4)

# Save
import os

BASE_DIR = "/Users/alexandrenuyt/Documents/Horizon-Énergie/Dimensionneur"
FILE_NAME = "Dimensionnement_Complet_v7_pro.xlsx"

path = os.path.join(BASE_DIR, FILE_NAME)
wb.save(path)
